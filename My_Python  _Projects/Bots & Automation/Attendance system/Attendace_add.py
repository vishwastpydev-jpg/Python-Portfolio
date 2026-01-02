import openpyxl
from openpyxl import Workbook
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import sys 
import getpass
import os

# --- EXCEL AUTO-CREATION LOGIC ---
def setup_excel(filename):
    if not os.path.exists(filename):
        print(f"📄 Creating new file: {filename}")
        wb = Workbook()
        sheet = wb.active
        sheet.title = "Sheet1"
        # New Header Structure: Roll, Name, Email, CI, Python, DM
        headers = ["Roll Number", "Student Name", "Email", "CI", "Python", "DM"]
        sheet.append(headers)
        wb.save(filename)
    else:
        print(f"📂 Using existing file: {filename}")

def add_students(filename):
    """Allows the user to input names and roll numbers into the sheet."""
    wb = openpyxl.load_workbook(filename)
    sheet = wb["Sheet1"]
    
    while True:
        print("\n--- Add New Student ---")
        roll = input("Enter Roll Number (or 'q' to finish): ")
        if roll.lower() == 'q': break
        
        name = input("Enter Student Name: ")
        email = input("Enter Student Email: ")
        
        # Append to Excel: Roll, Name, Email, CI=0, Python=0, DM=0
        sheet.append([int(roll), name, email, 0, 0, 0])
        wb.save(filename)
        print(f"✅ Added {name} (Roll: {roll})")

# --- CONFIGURATION ---
print("--- 📚 Attendance & Notification System ---")
EXCEL_PATH = input("Enter Excel filename (default: attendance.xlsx): ") or 'attendance.xlsx'
setup_excel(EXCEL_PATH)

# Ask to add students
manage = input("Do you want to add new students? (y/n): ").lower()
if manage == 'y':
    add_students(EXCEL_PATH)

SENDER_EMAIL = input("\nEnter your Gmail address: ")
SENDER_PASSWORD = getpass.getpass("Enter your Gmail App Password: ") 

staff_mails = [
    input("Enter CI Staff Email: "),
    input("Enter Python Staff Email: "),
    input("Enter DM Staff Email: ")
]

# --- LOAD UPDATED WORKBOOK ---
book = openpyxl.load_workbook(EXCEL_PATH)
sheet = book['Sheet1']
resp = 1

# --- HELPER FUNCTIONS ---
def savefile():
    book.save(EXCEL_PATH)

def mail_setup_and_send(recipients, subject, msg_body):
    if not recipients: return
    try:
        s = smtplib.SMTP('smtp.gmail.com', 587, timeout=60)
        s.starttls()
        s.login(SENDER_EMAIL, SENDER_PASSWORD)
        for recipient in recipients:
            message = MIMEMultipart()
            message['Subject'] = subject
            message.attach(MIMEText(msg_body, 'plain'))
            s.sendmail(SENDER_EMAIL, recipient, message.as_string())
        s.quit()
        print(f"✅ Emails sent to {len(recipients)} person(s).")
    except Exception as e:
        print(f"❌ Mail Error: {e}")

def check_attendance(no_of_days, row_num, sub_code):
    l1 = []; l2 = ""; l3 = []
    subjects = {1: "CI", 2: "Python", 3: "Data Mining"}
    sub_name = subjects.get(sub_code)

    for i in range(len(row_num)):
        leaves = no_of_days[i]
        row_idx = row_num[i]
        # Column 2 is Name, Column 3 is Email
        name = sheet.cell(row=row_idx, column=2).value
        email = sheet.cell(row=row_idx, column=3).value
        roll = sheet.cell(row=row_idx, column=1).value
        
        if leaves == 2:
            l1.append(email)
        elif leaves > 2:
            l2 += f"{name}({roll}), "
            l3.append(email)

    if l1:
        mail_setup_and_send(l1, f"Final Warning: {sub_name}", f"Dear Student, you have 2 leaves in {sub_name}.")
    if l3:
        student_msg = f"Lack of attendance in {sub_name}. Contact staff."
        staff_msg = f"Attendance Alert: The following students crossed leave limits in {sub_name}: {l2.strip(', ')}"
        mail_setup_and_send(l3, f"Lack of Attendance: {sub_name}", student_msg)
        mail_setup_and_send([staff_mails[sub_code-1]], f"Staff Report: {sub_name}", staff_msg)

# --- MAIN LOOP ---
while resp == 1:
    print("\nSelect Subject: 1->CI | 2->Python | 3->DM")
    try:
        sub_choice = int(input("Choice: "))
        raw_rolls = input("Enter roll numbers of absentees (separated by space): ")
        roll_list = list(map(int, raw_rolls.split()))

        row_indices = []
        current_leave_counts = []
        # Adjusted Column mapping: 4=CI, 5=Python, 6=DM
        sub_col = {1: 4, 2: 5, 3: 6}.get(sub_choice)

        for roll in roll_list:
            found = False
            for i in range(2, sheet.max_row + 1):
                if sheet.cell(row=i, column=1).value == roll:
                    val = sheet.cell(row=i, column=sub_col).value or 0
                    new_val = val + 1
                    sheet.cell(row=i, column=sub_col).value = new_val
                    row_indices.append(i)
                    current_leave_counts.append(new_val)
                    found = True
                    break
            if not found:
                print(f"⚠️ Roll {roll} not found.")

        savefile()
        if row_indices:
            check_attendance(current_leave_counts, row_indices, sub_choice)

        resp = int(input("\nAnother subject? (1 for Yes, 0 for No): "))
    except ValueError:
        print("Invalid input.")
        resp = 0

print("Attendance updated and saved.")
